import openpyxl as px
from openpyxl import load_workbook
import pandas as pd
from openpyxl.worksheet.table import Table, TableStyleInfo
from .protocols import Protocols

class ExcelConnector:
    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = self._load_file()
        self.ws = self.wb.active
        self.df = self._load_dataframe()
  

    def _load_file(self):
     
     try:
            wb = load_workbook(self.file_path)
            return wb
     except:
         wb = px.Workbook()
         wb.save(self.file_path)
         return wb
     
    def _load_dataframe(self):
        try:
            df = pd.read_excel(self.file_path, sheet_name=self.ws.title)
            return df
        except:
            df = pd.DataFrame()
            return df
        
    def save_dataframe(self):
        with pd.ExcelWriter(self.file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
          self.df.to_excel(writer, sheet_name=self.ws.title, index=False)

        return True
    def set_active_sheet(self, sheet_name):
        if sheet_name in self.wb.sheetnames:
            self.ws = self.wb[sheet_name]
            self.df = self._load_dataframe()
            return True
        else:
                print(f"Sheet '{sheet_name}' does not exist.")
                return False
        
    def create_sheet(self, sheet_name):
        try:
            self.wb.create_sheet(title=sheet_name)
            self.wb.save(self.file_path)
            self.ws = self.wb[sheet_name]
            self.df = pd.DataFrame()
            return True
        except:
            print(f"Failed to create sheet '{sheet_name}'.")
            return False
        
    def create_table(self, columns):

        try:
            self.ws.delete_rows(1, self.ws.max_row)
            self.ws.append(columns)
            self.df = pd.DataFrame(columns=columns)
            self.save_dataframe()
            return True
        except Exception as e:
            print(f"Failed to create table: {e}")
            return False
        
    def add_row(self, value_dict):
        try:
            new_value = pd.DataFrame([value_dict])
            self.df = pd.concat([self.df, new_value], ignore_index= True)
            self.save_dataframe()
            return self.get_preview()    
        except Exception as e:
            print(f"Failed to add row: {e}")
            return False
        
    
    def get_preview(self, n=5):
        try:
            self.df.head(n)
            preview = self.df.head(n).to_dict(orient='records')
            return ("status:success", preview)
        except Exception as e:
            return ("status:error, message:", str(e))
        
    def update_cell(self, column_name, row_index, new_value):
        try:
            if column_name in self.df.columns and 0<=row_index < len(self.df):
                self.df.loc[row_index, column_name] = new_value
                self.save_dataframe()
                return self.get_preview()
            else:
                print("Invalid column name or row index.")
                return False
        except Exception as e:
            print(f"failed to update cell:", e)
            return False
        
    def sum_column(self, column_name):
            """SUM only ONE column — per MVP spec"""
            try:
                if column_name not in self.df.columns:
                    return {
                        "status": "error",
                        "message": f"Column '{column_name}' not found."
                    }

                numeric = pd.to_numeric(self.df[column_name], errors='coerce')
                skipped = numeric.isna().sum()
                total = numeric.sum()

                return {
                    "status": "success",
                    "column": column_name,
                    "sum": float(total),
                    "skipped_values": int(skipped)
                }
            except Exception as e:
                return {"status": "error", "message": str(e)}
            
    def execute(self, action, params=None, mode=None):
        protocol = Protocols()

        if params is None:
            params = {}

        try:
            # 1. Validate action + execution intent
            protocol.validate_action(action, mode)

            # 2. Validate parameters strictly
            protocol.validate_params(action, params)

        except ValueError as e:
            return {
                "status": "error",
                "message": str(e)
            }

        actions = {
            "create_table": self.create_table,
            "add_row": self.add_row,
            "update_cell": self.update_cell,
            "sum_column": self.sum_column,
            "preview": self.get_preview,
            "set_active_sheet": self.set_active_sheet,
            "create_sheet": self.create_sheet
        }

        if action not in actions:
            return {
                "status": "error",
                "message": f"Action missing: {action}"
            }

        try:
            result = actions[action](**params)
            return {
                "status": "success",
                "data": result
            }

        except Exception as e:
            return {
                "status": "error",
                "message": str(e)
            }
